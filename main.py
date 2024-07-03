from openpyxl import Workbook
import src
import json


class Main:
    def generate_report(self) -> None:
        data = self.__get_data()
        workbook = Workbook()

        self.insert_customers_region(workbook, data['customers_region'])
        self.insert_customers_purchases(workbook, data['customers_purchases'])
        self.insert_employees_region(workbook, data['employees_region'])
        self.insert_employees_sales(workbook, data['employees_sales'])
        self.insert_employees_supervision(workbook, data['employees_supervision'])
        self.insert_product_most_sold(workbook, data['most_sold'])
        self.insert_product_most_profitable(workbook, data['most_profitable'])

        workbook.save('report.xlsx')

    def __get_data(self) -> dict:
        instances = src.initialize_classes()

        customers_region = instances['customer_info'].region()
        customers_purchases = instances['customer_info'].n_purchases()

        employees_region = instances['employees_info'].region()
        employees_purchases = instances['employees_info'].n_sales()
        employees_supervision = instances['employees_info'].supervision()

        most_sold = instances['sales_info'].products_solds('total_sold', 5)
        most_profitable = instances['sales_info'].products_solds('total_amount', 5)

        return {
            'customers_region': customers_region,
            'customers_purchases': customers_purchases,
            'employees_region': employees_region,
            'employees_sales': employees_purchases,
            'employees_supervision': employees_supervision,
            'most_sold': most_sold,
            'most_profitable': most_profitable
        }

    def insert_customers_region(self, workbook: Workbook, data: object) -> Workbook:
        sheet = workbook.active
        sheet.title = 'Customers Country'
        sheet['A1'] = 'Country'
        sheet['B1'] = 'Number of Customers'
        self.format_to_insert(sheet, data)

        return sheet

    def insert_customers_purchases(self, workbook: Workbook, data: object) -> Workbook:
        sheet = workbook.create_sheet(title='Customers Purchases')
        sheet['A1'] = 'Customers Id'
        sheet['B1'] = 'Company Name'
        sheet['C1'] = 'Purchases Number'
        sheet['D1'] = 'Contact Name'
        sheet['E1'] = 'Phone'
        self.format_to_insert(sheet, data)

        return sheet

    def insert_employees_region(self, workbook: Workbook, data: object) -> Workbook:
        sheet = workbook.create_sheet(title='Employees Country')
        sheet['A1'] = 'Country'
        sheet['B1'] = 'Number of Employees'
        self.format_to_insert(sheet, data)

        return sheet

    def insert_employees_sales(self, workbook: Workbook, data: object) -> Workbook:
        sheet = workbook.create_sheet(title='Employees Sales')
        sheet['A1'] = 'Employee Id'
        sheet['B1'] = 'Employee Name'
        sheet['C1'] = 'Number of Sales'
        self.format_to_insert(sheet, data)

        return sheet

    def insert_employees_supervision(self, workbook: Workbook, data: object) -> Workbook:
        sheet = workbook.create_sheet(title='Supervisors')
        sheet['A1'] = 'Supervisor Name'
        sheet['B1'] = 'Supervises List'
        sheet['C1'] = 'Number of Supervised'
        self.format_to_insert(sheet, data)

        return sheet

    def insert_product_most_sold(self, workbook: Workbook, data: object) -> Workbook:
        sheet = workbook.create_sheet(title='Product Most Sold')
        sheet['A1'] = 'Product Id'
        sheet['B1'] = 'Product Name'
        sheet['C1'] = 'Total Sold'
        sheet['D1'] = 'Total Amount'
        self.format_to_insert(sheet, data)

        return sheet

    def insert_product_most_profitable(self, workbook: Workbook, data: object) -> Workbook:
        sheet = workbook.create_sheet(title='Product Most Profitable')
        sheet['A1'] = 'Product Id'
        sheet['B1'] = 'Product Name'
        sheet['C1'] = 'Total Sold'
        sheet['D1'] = 'Total Amount'
        self.format_to_insert(sheet, data)

        return sheet

    def format_to_insert(self, sheet: Workbook, data) -> Workbook:
        insert_row = 2

        if isinstance(data, dict):
            for key, value in data.items():
                sheet.cell(row=insert_row, column=1, value=key)
                sheet.cell(row=insert_row, column=2, value=value)
                insert_row += 1

        elif isinstance(data, list):
            for each in data:
                for col, value in enumerate(each.values(), start=1):
                    self.confirm_insert_type(sheet, value, insert_row, col)
                insert_row += 1

        return sheet

    def confirm_insert_type(self, sheet: Workbook, value, insert_row: int, col: int) -> Workbook:
        if isinstance(value, list):
            item_list = ''
            for item in value:
                item_list += json.dumps(item, indent=4)
            sheet.cell(row=insert_row, column=col, value=item_list)
        else:
            sheet.cell(row=insert_row, column=col, value=value)

        return sheet


Main().generate_report()
